"""
Shared export utility — Excel, PDF, and email export for reports and tables.
"""

import io
import logging

logger = logging.getLogger(__name__)


def export_table_to_excel(headers: list, rows: list, sheet_name: str = "Report") -> bytes:
    """Export tabular data to Excel (.xlsx) bytes using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_html_to_pdf(html_content: str) -> bytes:
    """Export HTML content to PDF bytes using weasyprint."""
    try:
        from weasyprint import HTML
    except ImportError:
        logger.error("weasyprint not installed — run: pip install weasyprint")
        return b""

    buf = io.BytesIO()
    HTML(string=html_content).write_pdf(buf)
    return buf.getvalue()


def send_report_email(to: str, subject: str, body: str,
                      attachment_bytes: bytes = None, attachment_name: str = None) -> bool:
    """Send report via email using Postmark (postmarker)."""
    import os
    try:
        from postmarker.core import PostmarkClient
    except ImportError:
        logger.error("postmarker not installed — run: pip install postmarker")
        return False

    api_key = os.getenv("POSTMARK_API_KEY")
    from_email = os.getenv("POSTMARK_FROM_EMAIL", "reports@ashland-hill.com")
    if not api_key:
        logger.error("POSTMARK_API_KEY not set")
        return False

    client = PostmarkClient(server_token=api_key)
    kwargs = {
        "From": from_email,
        "To": to,
        "Subject": subject,
        "HtmlBody": body,
    }
    if attachment_bytes and attachment_name:
        import base64
        content_type = "application/pdf" if attachment_name.endswith(".pdf") else \
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        kwargs["Attachments"] = [{
            "Name": attachment_name,
            "Content": base64.b64encode(attachment_bytes).decode(),
            "ContentType": content_type,
        }]

    try:
        client.emails.send(**kwargs)
        return True
    except Exception as e:
        logger.error(f"Email send failed: {e}")
        return False
